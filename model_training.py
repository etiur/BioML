from sklearn.model_selection import StratifiedShuffleSplit
import pandas as pd
from sklearn.metrics import matthews_corrcoef, confusion_matrix, r2_score
from sklearn.metrics import classification_report as class_re
from collections import namedtuple
from openpyxl import load_workbook
from hpsklearn import HyperoptEstimator
from utilities import scale, write_excel
from hyperopt import hp, tpe
from pathlib import Path
from hpsklearn import random_forest_classifier, extra_trees_classifier
from hpsklearn import sgd_classifier, ridge_classifier, passive_aggressive_classifier
from hpsklearn import mlp_classifier, k_neighbors_classifier, xgboost_classification, lightgbm_classification
from hpsklearn import svc
import argparse
from concurrent.futures import ProcessPoolExecutor as Pool


def arg_parse():
    parser = argparse.ArgumentParser(description="Train the models")

    parser.add_argument("-o", "--training_output", required=False,
                        help="The path where to save the models training results",
                        default="training_results")
    parser.add_argument("-l", "--label", required=True,
                        help="The path to the labels of the training set in a csv format")
    parser.add_argument("-n", "--num_thread", required=False, default=50, type=int,
                        help="The number of threads to search for the hyperparameter space")
    parser.add_argument("-s", "--scaler", required=False, default="robust", choices=("robust", "standard", "minmax"),
                        help="Choose one of the scaler available in scikit-learn, defaults to RobustScaler")
    parser.add_argument("-e", "--excel", required=False,
                        help="The file to where the selected features are saved in excel format",
                        default="training_features/selected_features.xlsx")
    parser.add_argument("-k", "--kfold_parameters", required=False,
                        help="The parameters for the kfold in num_split:test_size format", default="5:0.2")
    parser.add_argument("-ot", "--outliers", nargs="+", required=False, default=(),
                        help="A list of outliers if any, the name should be the same as in the excel file with the "
                             "filtered features, you can also specify the path to a file in plain text format, each "
                             "record should be in a new line")
    parser.add_argument("-hp", "--hyperparameter_tuning", required=False, default="80:30",
                        help="The parameters for the class that performs hyperparameter tuning"
                             " in max_evals: trial_time format")
    parser.add_argument("-pw", "--precision_weight", required=False, default=1, type=float,
                        help="Weights to specify how relevant is the precision for the ranking of the different "
                             "features")
    parser.add_argument("-rw", "--recall_weight", required=False, default=0.8, type=float,
                        help="Weights to specify how relevant is the recall for the ranking of the different features")
    parser.add_argument("-c0", "--class0_weight", required=False, default=0.5, type=float,
                        help="Weights to specify how relevant is the f1, precision and recall scores of the class 0"
                             " or the negative class for the ranking of the different features with respect to class 1 "
                             "or the positive class")
    parser.add_argument("-rpw", "--report_weight", required=False, default=0.25, type=float,
                        help="Weights to specify how relevant is the f1, precision and recall for the ranking of the "
                             "different features with respect to MCC and the R2 which are more general measures of "
                             "the performance of a model")
    parser.add_argument("-dw", "--difference_weight", required=False, default=0.8, type=float,
                        help="How important is to have similar training and test metrics")

    args = parser.parse_args()

    return [args.label, args.training_output, args.hyperparameter_tuning, args.num_thread, args.scaler,
            args.excel, args.kfold_parameters, args.outliers, args.precision_weight, args.recall_weight,
            args.class0_weight, args.report_weight, args.difference_weight]


def interesting_classifiers(name):
    """
    All classifiers
    """
    classifiers = [
        random_forest_classifier(name + ".random_forest"),
        extra_trees_classifier(name + ".extra_trees"),
        sgd_classifier(name + ".sgd"),
        ridge_classifier(name + ".ridge", random_state=0),
        passive_aggressive_classifier(name + ".passive_aggressive"),
        mlp_classifier(name + ".mlp"),
        svc(name + ".svc"),
        xgboost_classification(name + ".xgboost"),
        lightgbm_classification(name + ".lightgbm"),
        k_neighbors_classifier(name + ".knn")
    ]

    return hp.choice(name, classifiers)


class Classifier:
    def __init__(self, feature_path, label, training_output="training_results", num_splits=5, test_size=0.2,
                 outliers=(), scaler="robust", max_evals=200, trial_time=30, num_threads=10, precision_weight=1,
                 recall_weight=1, report_weight=0.4, difference_weight=0.8, class0_weight=0.5):
        self.outliers = outliers
        self.num_splits = num_splits
        self.test_size = test_size
        self.features = Path(feature_path)  # if only one features it will perform split and train
        self.labels = pd.read_csv(label, index_col=0)
        self.scaler = scaler
        self.max_evals = max_evals
        self.trial_time_out = trial_time
        self.num_threads = num_threads
        self.output_path = Path(training_output)  # for the model results
        self.output_path.mkdir(parents=True, exist_ok=True)
        self.pre_weight = precision_weight
        self.rec_weight = recall_weight
        self.report_weight = report_weight
        self.difference_weight = difference_weight
        self.class0_weight = class0_weight
        self.with_split = True

    def train(self, X_train, Y_train, X_test):
        estimator = HyperoptEstimator(classifier=interesting_classifiers("automl"), preprocessing=[], algo=tpe.suggest,
                                      max_evals=self.max_evals, trial_timeout=self.trial_time_out,
                                      n_jobs=self.num_threads, verbose=True)
        estimator.fit(X_train, Y_train)
        # Model predictions
        pred_test_y = estimator.predict(X_test)
        # training data prediction
        pred_train_y = estimator.predict(X_train)
        pred = namedtuple("prediction", ["fitted", "pred_test_y", "pred_train_y"])
        return pred(*[estimator, pred_test_y, pred_train_y])

    @staticmethod
    def get_score(pred, Y_train, Y_test):
        """ The function prints the scores of the models and the prediction performance """
        target_names = ["class 0", "class 1"]

        scalar_record = namedtuple("scores", ["cv_score", "train_mat", "test_matthews", "r2_train", "r2_test"])
        parameter_record = namedtuple("parameters", ["params", "test_confusion", "tr_report", "te_report",
                                                     "train_confusion", "model_name"])
        # Model comparison
        cv_score = 1 - pred.fitted._best_loss
        model_params = pred.fitted.best_model()["learner"].get_params()
        model_params = {key: value for key, value in model_params.items() if key not in ["warm_start", "verbose",
                                                                                         "oob_score"]}
        model_params = pd.Series(model_params)
        # Training scores
        train_confusion = confusion_matrix(Y_train, pred.pred_train_y)
        tr_report = class_re(Y_train, pred.pred_train_y, target_names=target_names, output_dict=True)
        train_mat = matthews_corrcoef(Y_train, pred.pred_train_y)
        train_r2 = r2_score(Y_train, pred.pred_train_y)
        # Test metrics grid
        test_confusion = confusion_matrix(Y_test, pred.pred_test_y)
        test_matthews = matthews_corrcoef(Y_test, pred.pred_test_y)
        te_report = class_re(Y_test, pred.pred_test_y, target_names=target_names, output_dict=True)
        test_r2 = r2_score(Y_test, pred.pred_test_y)
        # save the results in namedtuples
        scalar_scores = scalar_record(*[cv_score, train_mat, test_matthews, train_r2, test_r2])
        param_scores = parameter_record(*[model_params, test_confusion, tr_report, te_report, train_confusion,
                                          pred.fitted.best_model()["learner"].__class__.__name__])

        return scalar_scores, param_scores

    def _scale(self, features, train_index, test_index, split_index):
        # split and filter
        if self.with_split:
            feat_subset = features.loc[:, f"split_{split_index}"] # each split different features and different fold of
            # training and test data
        else:
            feat_subset = features
        X_train = feat_subset.iloc[train_index]
        Y_train = feat_subset.iloc[train_index]
        X_test = feat_subset.iloc[test_index]
        Y_test = feat_subset.iloc[test_index]
        X_train = X_train.loc[[x for x in X_train.index if x not in self.outliers]]
        X_test = X_test.loc[[x for x in X_test.index if x not in self.outliers]]
        Y_train = Y_train.loc[[x for x in Y_train.index if x not in self.outliers]].values.ravel()
        Y_test = Y_test.loc[[x for x in Y_test.index if x not in self.outliers]].values.ravel()
        transformed_x, scaler_dict, test_x = scale(self.scaler, X_train, X_test)

        return transformed_x, test_x, Y_test, Y_train

    def _check_features(self, sheet_names):
        features = pd.read_excel(self.features, index_col=0, header=[0, 1], engine='openpyxl')
        if f"split_{0}" not in features.columns.unique(level=0):
            self.with_split = False
        if self.with_split:
            excel_data = pd.read_excel(self.features, index_col=0, sheet_name=sheet_names, header=[0, 1])
        else:
            excel_data = pd.read_excel(self.features, index_col=0, sheet_name=sheet_names, header=0)
        return excel_data

    def nested_cv(self, feature):
        """Performs something similar to a nested cross-validation"""
        metric_scalar = []
        parameter_list = []
        split_index = []
        skf = StratifiedShuffleSplit(n_splits=self.num_splits, test_size=self.test_size, random_state=20)
        for ind, (train_index, test_index) in enumerate(skf.split(feature, self.labels)):
            split_index.append(ind)
            transformed_x, test_x, Y_test, Y_train = self._scale(feature, train_index, test_index, ind)
            pred = self.train(transformed_x, Y_train, test_x)
            scalar_scores, param_scores = self.get_score(pred, Y_train, Y_test)
            metric_scalar.append(scalar_scores)
            parameter_list.append(param_scores)

        return metric_scalar, parameter_list, split_index

    @staticmethod
    def to_dataframe(metric_scalar, parameter_list, split_index):
        matrix = namedtuple("confusion_matrix", ["true_n", "false_p", "false_n", "true_p"])
        # performance scores
        test_mathew = [x.test_matthews for x in metric_scalar]
        cv_score = [x.cv_score for x in metric_scalar]
        test_r2 = [x.r2_test for x in metric_scalar]
        train_mathew = [x.train_mat for x in metric_scalar]
        train_r2 = [x.r2_train for x in metric_scalar]
        # model parameters
        model_name = [x.model_name for x in parameter_list]
        params = pd.concat({i: pd.concat({x.model_name: x.params}) for i, x in zip(split_index, parameter_list)})
        # Taking the confusion matrix
        test_confusion = [matrix(*x.test_confusion.ravel()) for x in parameter_list]
        training_confusion = [matrix(*x.train_confusion.ravel()) for x in parameter_list]
        te_report = {num: pd.DataFrame(x.te_report).transpose() for num, x in zip(split_index, parameter_list)}
        te_report = pd.concat(te_report)
        te_report.index.names = ["split_index", "labels"]
        tr_report = {num: pd.DataFrame(x.tr_report).transpose() for num, x in zip(split_index, parameter_list)}
        tr_report = pd.concat(tr_report)
        tr_report.index.names = ["split_index", "labels"]
        report = pd.concat({"train": tr_report.T, "test": te_report.T})
        # Separating confusion matrix into individual elements
        test_true_n = [y.true_n for y in test_confusion]
        test_false_p = [y.false_p for y in test_confusion]
        test_false_n = [y.false_n for y in test_confusion]
        test_true_p = [y.true_p for y in test_confusion]

        training_true_n = [z.true_n for z in training_confusion]
        training_false_p = [z.false_p for z in training_confusion]
        training_false_n = [z.false_n for z in training_confusion]
        training_true_p = [z.true_p for z in training_confusion]

        dataframe = pd.DataFrame([split_index, test_true_n, test_true_p, test_false_p, test_false_n, training_true_n,
                                  training_true_p, training_false_p, training_false_n, cv_score,
                                  train_mathew, test_mathew, train_r2, test_r2, model_name])

        dataframe = dataframe.transpose()
        dataframe.columns = ["split_index", "test_tn", "test_tp", "test_fp", "test_fn", "train_tn", "train_tp",
                             "train_fp", "train_fn", "CV_MCC", "train_MCC", "test_MCC", "train_R2", "test_R2",
                             "model_name"]
        dataframe.set_index("split_index", inplace=True)

        return dataframe, report.T, params

    def _calculate_score_dataframe(self, dataframe):
        return ((dataframe["train_MCC"] + dataframe["test_MCC"] + dataframe["train_R2"] + dataframe["test_R2"])
                - self.difference_weight * (abs(dataframe["test_MCC"] - dataframe["train_MCC"]) +
                                            abs(dataframe["test_R2"] - dataframe["train_R2"]))).sum()

    def _calculate_score_report(self, report, class_label):
        class_level = report.loc[report.index.get_level_values(1) == class_label,
                                 report.columns.get_level_values(1).isin(["precision", "recall"])]
        pre_sum = self.pre_weight * (class_level.loc[(slice(None), class_label), ("test", "precision")] +
                                     class_level.loc[(slice(None), class_label), ("train", "precision")])
        rec_sum = self.rec_weight * (class_level.loc[(slice(None), class_label), ("test", "recall")] +
                                     class_level.loc[(slice(None), class_label), ("train", "recall")])
        pre_diff = self.pre_weight * abs(class_level.loc[(slice(None), class_label), ("test", "precision")] -
                                         class_level.loc[(slice(None), class_label), ("train", "precision")])
        rec_diff = self.rec_weight * abs(class_level.loc[(slice(None), class_label), ("test", "recall")] -
                                         class_level.loc[(slice(None), class_label), ("train", "recall")])
        return (pre_sum + rec_sum - self.difference_weight * (pre_diff + rec_diff)).sum()

    def rank_results(self, result_item: list) -> float:

        sheet, dataframe, report, params = result_item

        score_dataframe = self._calculate_score_dataframe(dataframe)
        score_report_class0 = self._calculate_score_report(report, "class 0")
        score_report_class1 = self._calculate_score_report(report, "class 1")
        # Calculate the overall rank as the sum of the scores -> the greater, the better
        rank = score_dataframe + (self.report_weight * (self.class0_weight * score_report_class0 +
                                                        score_report_class1)/2)

        return rank

    def _run(self, feature):
        # creating list of results
        metric_scalar, parameter_list, split_index = self.nested_cv(feature)
        dataframe, report, params = self.to_dataframe(metric_scalar, parameter_list, split_index)
        return dataframe, report, params

    def run(self):
        """A function that runs nested_cv several times, as many as the sheets in the Excel"""
        # reading the data
        sheet_names = load_workbook(self.features).sheetnames
        excel_data = self._check_features(sheet_names)
        result_list = []
        with Pool(self.num_threads) as pool:
            for num, (dataframe, report, params) in enumerate(pool.map(self._run, excel_data.values())):
                result_list.append((sheet_names[num], dataframe, report, params))

        result_list.sort(key=self.rank_results, reverse=True)
        with (pd.ExcelWriter(self.output_path/"training_results.xlsx", mode="w", engine="openpyxl") as writer1,
              pd.ExcelWriter(self.output_path/"hyperparameters.xlsx", mode="w", engine="openpyxl") as writer2,
              pd.ExcelWriter(self.output_path/"classification_report.xlsx", mode="w", engine="openpyxl") as writer3):
            for x in result_list:
                sheet, dataframe, report, params = x
                write_excel(writer1, dataframe, sheet)
                write_excel(writer2, params, sheet)
                write_excel(writer3, report, sheet)


def main():
    label, training_output, hyperparameter_tuning, num_thread, scaler, excel, kfold, outliers, \
        precision_weight, recall_weight, class0_weight, report_weight, difference_weight = arg_parse()
    num_split, test_size = int(kfold.split(":")[0]), float(kfold.split(":")[1])
    max_evals, trial_time = int(hyperparameter_tuning.split(":")[0]), int(hyperparameter_tuning.split(":")[1])
    if Path(outliers[0]).exists():
        with open(outliers) as out:
            outliers = [x.strip() for x in out.readlines()]
    training = Classifier(excel, label, training_output, num_split, test_size, outliers, scaler, max_evals,
                          trial_time, num_thread, precision_weight, recall_weight, report_weight, difference_weight,
                          class0_weight)
    training.run()


if __name__ == "__main__":
    # Run this if this file is executed from command line but not if is imported as API
    main()
