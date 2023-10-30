from pyod.models.abod import ABOD
# from pyod.models.cblof import CBLOF
from pyod.models.feature_bagging import FeatureBagging
from pyod.models.hbos import HBOS
from pyod.models.iforest import IForest
from pyod.models.knn import KNN
from pyod.models.lof import LOF
from pyod.models.pca import PCA
from pyod.models.ocsvm import OCSVM
from pyod.models.ecod import ECOD
from openpyxl import load_workbook
import pandas as pd
from .utilities import scale
from pathlib import Path
import random
import argparse


def arg_parse():
    parser = argparse.ArgumentParser(description="Detect outliers from the selected features")

    parser.add_argument("-e", "--excel", required=False,
                        help="The file to where the selected features are saved in excel format",
                        default="training_features/selected_features.xlsx")
    parser.add_argument("-o", "--outlier", required=False,
                        help="The path to the output for the outliers",
                        default="training_results/outliers.csv")
    parser.add_argument("-n", "--num_thread", required=False, default=10, type=int,
                        help="The number of threads to use for the parallelization of outlier detection")
    parser.add_argument("-s", "--scaler", required=False, default="robust", choices=("robust", "standard", "minmax"),
                        help="Choose one of the scaler available in scikit-learn, defaults to RobustScaler")
    parser.add_argument("-c", "--contamination", required=False, default=0.06, type=float,
                        help="The expected % of outliers")
    parser.add_argument("-nfe", "--num_features", required=False, type=float,
                        help="The fraction of features to use, maximum 1 which is all the features", default=0.8)

    args = parser.parse_args()

    return [args.excel, args.outlier, args.scaler, args.contamination, args.num_thread, args.num_features]


class OutlierDetection:
    def __init__(self, feature_file="training_features/selected_features.xlsx", output="training_results/outliers.csv",
                 scaler="robust", contamination=0.06, num_thread=10, num_feature=1):
        self.scaler = scaler
        self.contamination = contamination
        self.feature_file = feature_file
        self.num_threads = num_thread
        self.book = load_workbook(feature_file, read_only=True)
        self.output = Path(output)
        self.output.parent.mkdir(parents=True, exist_ok=True)
        self.num_feature = num_feature
        if not str(self.output).endswith(".csv"):
            self.output.with_suffix(".csv")
        self.with_split = True

    def outlier(self, transformed_x):
        """Given a model it will return all its scores for each of the worksheets"""

        iforest = IForest(n_estimators=200, random_state=0, max_features=0.8, contamination=self.contamination,
                          n_jobs=self.num_threads)
        knn = KNN(method="mean", contamination=self.contamination, n_jobs=self.num_threads)
        bagging = FeatureBagging(LOF(), random_state=20, contamination=self.contamination, n_jobs=self.num_threads)
        # cblof = CBLOF(random_state=10)
        hbos = HBOS(contamination=self.contamination)
        abod = ABOD(contamination=self.contamination)
        pca = PCA(contamination=self.contamination)
        ocsvm = OCSVM(contamination=self.contamination)
        ecod = ECOD(contamination=self.contamination, n_jobs=self.num_threads)
        classifiers = {"iforest": iforest, "knn": knn, "bagging": bagging, "hbos": hbos, "abod": abod,
                       "pca": pca, "ocsvm": ocsvm, "ecod": ecod}

        prediction = {}
        raw = {}
        for name, clf in classifiers.items():
            # for the iforest
            clf.fit(transformed_x)
            train_pred = clf.labels_  # binary labels (0: inliers, 1: outliers)
            train_scores = clf.decision_scores_  # raw outlier scores
            prediction[name] = train_pred
            raw[name] = train_scores

        return prediction

    @staticmethod
    def counting(prediction, index):

        pred = {key: pd.DataFrame(value).T for key, value in prediction.items()}
        pred = pd.concat(pred)
        summed = pred.sum()
        summed.index = index
        summed.sort_values(ascending=False, inplace=True)
        return summed

    def _check_features(self, book):
        features = pd.read_excel(self.feature_file, index_col=0, header=[0, 1], engine='openpyxl')
        if f"split_{0}" not in features.columns.unique(0):
            self.with_split = False
        if self.with_split:
            new_excel = {}
            excel_data = pd.read_excel(self.feature_file, index_col=0, sheet_name=book, header=[0, 1])
            for key, values in excel_data.items():
                for col in values.columns.unique(level=0):
                    ind = int(col.split("_")[-1])
                    new_excel[f"{key}_{ind}"] = values.loc[:, f"split_{ind}"]
            excel_data = new_excel
        else:
            excel_data = pd.read_excel(self.feature_file, index_col=0, sheet_name=book, header=0)
        return excel_data

    def run(self):
        results = {}
        book = self.book.sheetnames
        excel_data = self._check_features(book)
        excel_data = {key: x.sample(frac=1, random_state=0) for key, x in excel_data.items()}
        scaled_data = []
        for key, x in excel_data.items():
            transformed_x, scaler_dict = scale(self.scaler, x)
            scaled_data.append((key, transformed_x))
        # parallelized
        scaled_data = random.sample(scaled_data, int(len(scaled_data)*self.num_feature))
        for key, scaled in scaled_data:
            print(f"using {key} for outlier calculations")
            res = self.outlier(scaled)
            results[key] = res

        summed = self.counting(results, x.index)
        print("saving the outlier file")
        summed.to_csv(self.output)
        return summed


def main():
    excel, outlier, scaler, contamination, num_thread, num_features = arg_parse()
    detection = OutlierDetection(excel, outlier, scaler, contamination, num_thread, num_features)
    detection.run()


if __name__ == "__main__":
    # Run this if this file is executed from command line but not if is imported as API
    main()

